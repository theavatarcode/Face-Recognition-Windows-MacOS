import shutil as s
import PySimpleGUI as sg
import os
import cv2
import webbrowser as wb
import numpy as np
import face_recognition
import glob
import datetime as dt
import openpyxl as xl

class SimpleFacerec:
    def __init__(self):
        self.known_face_encodings = []
        self.known_face_names = []  
        self.frame_resizing = 0.25
    def load_encoding_images(self, images_path):
        
        images_path = glob.glob(os.path.join(images_path, "*.*"))
        print("{} encoding images found.".format(len(images_path)))
        for img_path in images_path:
            img = cv2.imread(img_path)
            rgb_img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
            basename = os.path.basename(img_path)
            (filename, ext) = os.path.splitext(basename)     
            img_encoding = face_recognition.face_encodings(rgb_img)[0]    
            self.known_face_encodings.append(img_encoding)
            self.known_face_names.append(filename)
        print("Encoding images loaded")

    def detect_known_faces(self, frame):
        small_frame = cv2.resize(frame, (0, 0), fx=self.frame_resizing, fy=self.frame_resizing)
        
        rgb_small_frame = cv2.cvtColor(small_frame, cv2.COLOR_BGR2RGB)
        face_locations = face_recognition.face_locations(rgb_small_frame)
        face_encodings = face_recognition.face_encodings(rgb_small_frame, face_locations)
        face_percent_value = 0
        percent = 0
        face_names = []
        for face_encoding in face_encodings:           
            matches = face_recognition.compare_faces(self.known_face_encodings, face_encoding)
            name = "Unknown"
            face_distances = face_recognition.face_distance(self.known_face_encodings, face_encoding)
            best_match_index = np.argmin(face_distances)
            face_percent_value = (1-face_distances[best_match_index])*100
            percent = round(face_percent_value,2)
            
            if matches[best_match_index] and percent >=50 :
                name = self.known_face_names[best_match_index]
            else:
                name = 'Unknown'
                
            face_names.append(name)
            
        face_locations = np.array(face_locations)
        face_locations = face_locations / self.frame_resizing
        return face_locations.astype(int), face_names, percent

    def getClassesFrom(self, name):
        classes = str(name[0:3])
        return classes
    
    def getNumberFrom(self, n):
        number = str(n[4:6])
        return number
    
    def getNameFrom(self, n):
        onlyName = str(n[7:])
        return onlyName

def start():
    workbook = xl.Workbook()
    date = dt.datetime.now()
    dtt = date.strftime("%Y_%b_%d_%a %H-%M-%S")
    sheet = workbook.active
    sheet.title = 'Attendance'
    sheet["A1"] = "Name"
    sheet["B1"] = "ID"
    sheet["C1"] = "Class"
    sheet["D1"] = "Percent"
    sheet["E1"] = "Date/Time"



    nameincol = []
    row_start = sheet.max_row + 1
    sfr = SimpleFacerec()
    sfr.load_encoding_images('images/')
    cap = cv2.VideoCapture(0)
    process = True

    while True:
        
        ret, frame = cap.read()
        dt_now = str(dt.datetime.now())  
        face_locate = []
        
        
        if ret:
            face_locate, face_names, face_percent = sfr.detect_known_faces(frame)
            
            for (top,right,bottom,left), name in zip(face_locate, face_names):        
                    
                onlyname = sfr.getNameFrom(name)
                classes = sfr.getClassesFrom(name)
                id = sfr.getNumberFrom(name)
                
                
                if name =='Unknown':
                    onlyname = 'Unknown'
                    classes = 'Unknown'
                    id = 'Unknown'
                if onlyname not in nameincol and onlyname !='Unknown':
                    nameincol.append(onlyname)
                    sheet.cell(row= row_start, column= 1, value= onlyname)
                    sheet.cell(row= row_start, column= 2, value= id)
                    sheet.cell(row= row_start, column= 3, value= classes)
                    sheet.cell(row= row_start, column= 4, value= face_percent)
                    sheet.cell(row= row_start, column= 5, value= dt_now)
                    row_start += 1
                    
                        
                
                if name == "Unknown":
                    color = [46,2,209]
                else : 
                    color = [255,102,45]              
                cv2.rectangle(frame, (left,top), (right,bottom), color, 4)
                cv2.rectangle(frame, (left-1, top-30), (right+1,top), color, cv2.FILLED)
                font = cv2.FONT_HERSHEY_DUPLEX
                cv2.putText(frame, onlyname, (left+6,top-6), font, 0.8, (255,255,255), 1)
                cv2.putText(frame, str(face_percent) + "%", (left+6,bottom+26), font, 0.6, (255,255,255), 1)
            
        
        cv2.imshow("FaceRecognition by M6/3", frame)
        if cv2.waitKey(1) & 0xFF == ord("q"):
            break
    excel_path = os.path.realpath('ExcelData')
    workbook.save(f'{excel_path}/{dtt}.xlsx')

    cv2.destroyAllWindows()


#########################################################-GUI ZONE-#############################################################################################
sg.theme('LightGrey1')
font = ("Verdana", 14)
layout = [  [sg.Text('Add images to system. File must be only .JPG',font= font), sg.Button('Capture from Webcam',font=font)],
            [sg.Text('Path of Images',font=font), sg.InputText(key='-in-', do_not_clear=False,font=font), sg.FileBrowse(file_types= (("image","*.JPG*"),),font=font), sg.Button('Open',font=font)],
            [sg.Text('Name',font=font),sg.InputText(key = '-name-',size=(20,1), do_not_clear= False,font=font), sg.Text('ID',font=font), sg.InputText(key='-id-',size= (3,1), do_not_clear= False,font=font), sg.Text('Classroom (Ex. 6-3, 6-2, 6-1)',font=font), sg.InputText(key='-class-',size=(10,1),do_not_clear= False,font=font)],
            [sg.Text('Register face to system.',font=font),sg.Button('Upload',font=font)],
            [sg.VerticalSeparator(color=(0,0,0))],
            [sg.HorizontalSeparator(color=(0,0,0))],
            [sg.VerticalSeparator(color=(0,0,0))],
            [sg.StatusBar('Warming : Start Face-Recognition will start Excel attendance list too. ',key='-l-'),sg.Button('Start Face-Recognition',font=font), sg.Button('Cancel',font=font)],
            [sg.Text('Attendances list have save in Excel, You can open files here',font=font), sg.Button('OpenFolderExcel',font=font),sg.StatusBar('Support : Window7 64-bit, Window10 64-bit, MacOS')]]

icon = os.path.realpath('icon/ip-camera.ico')
excel = 'ExcelData' ## Path Fix!!!!
real_path_excelData = os.path.realpath(excel)
image_folder = os.path.realpath('images')


window = sg.Window('Face-Recogintion by M6/3', layout, icon= icon)
while True:
    event, values = window.read()
    if event == sg.WIN_CLOSED or event == 'Cancel': 
        break

    if event == 'Upload':
        name = str(values['-name-'])
        if len(str(values['-id-'])) == 1:
            id = '0'+str(values['-id-'])
        else:
            id = str(values['-id-'])
        classes = str(values['-class-'])
        old_name = values['-in-']
        try:
            
            locate_of_image = f"{image_folder}/{classes}_{id}_{name}.jpg" ## Path Fix!!!
            s.move(old_name, locate_of_image)
            sg.popup(f'image uploaded! move to {locate_of_image}')
        except:
            sg.popup('Error : None image or None info!, Path failed!')

    
    if event == 'Start Face-Recognition':
        sg.popup('Wait for endcoding images....')
        try:
            start()
        except:
            sg.popup('Error to Run!')


    if event == 'OpenFolderExcel':
        os.system(f'start {os.path.realpath(real_path_excelData)}')

    if event == 'Capture from Webcam':
   
        cap = cv2.VideoCapture(0)
        while True:
            ret, frame = cap.read()
            cv2.imshow('Press Spacebar to capture/ESC to quit', frame)
            k = cv2.waitKey(1)
            if k%256==27:
                break
            elif k%256 ==32:
                img_name = 'class_id_name.jpg'
                cv2.imwrite(img_name, frame)

                # path_file = f'C:/Users/PC/Desktop/face_rq/FaceRecognition/{img_name}'
                real_path_file_capture_image = os.path.realpath(img_name)
                window.FindElement('-in-').update(real_path_file_capture_image)
                values['-in-'] = real_path_file_capture_image
                break            
        cap.release()
        cv2.destroyAllWindows()
    
    if event =='Open':
        
        os.system(f'start {os.path.realpath(image_folder)}')

window.close()
