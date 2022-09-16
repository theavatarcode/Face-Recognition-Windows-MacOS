import cv2
from cv2 import VideoCapture
from simple_facerec import SimpleFacerec
import time
import datetime as dt
import openpyxl as xl
import os


def start():
    workbook = xl.Workbook()
    date = dt.datetime.now()
    dtt = date.strftime("%Y_%b_%d_%a___%H-%M-%S")
    sheet = workbook.active
    sheet.title = 'Attendance'
    sheet["A1"] = "Name"
    sheet["B1"] = "ID"
    sheet["C1"] = "Class"
    sheet["D1"] = "Percent"
    sheet["E1"] = "Date/Time"



    nameincol = []
    row_start = sheet.max_row + 1
    sfr = SimpleFacerec()
    sfr.load_encoding_images('images/')
    cap = cv2.VideoCapture(0)
    process = True

    while True:
        
        ret, frame = cap.read()
        realtimedate = dt.datetime.now()
        dt_now = str(realtimedate.strftime("%Y-"))
        
        face_locate = []
        
        
        if ret:
            face_locate, face_names, face_percent = sfr.detect_known_faces(frame)
            
            for (top,right,bottom,left), name in zip(face_locate, face_names):        
                    
                onlyname = sfr.getNameFrom(name)
                classes = sfr.getClassesFrom(name)
                id = sfr.getNumberFrom(name)
                
                
                if name =='Unknown':
                    onlyname = 'Unknown'
                    classes = 'Unknown'
                    id = 'Unknown'
                if onlyname not in nameincol and onlyname !='Unknown':
                    nameincol.append(onlyname)
                    sheet.cell(row= row_start, column= 1, value= onlyname)
                    sheet.cell(row= row_start, column= 2, value= id)
                    sheet.cell(row= row_start, column= 3, value= classes)
                    sheet.cell(row= row_start, column= 4, value= face_percent)
                    sheet.cell(row= row_start, column= 5, value= dt_now)
                    row_start += 1
                    
                        
                
                if name == "Unknown":
                    color = [46,2,209]
                else : 
                    color = [255,102,45]              
                cv2.rectangle(frame, (left,top), (right,bottom), color, 4)
                cv2.rectangle(frame, (left-1, top-30), (right+1,top), color, cv2.FILLED)
                font = cv2.FONT_HERSHEY_DUPLEX
                cv2.putText(frame, onlyname, (left+6,top-6), font, 0.8, (255,255,255), 1)
                cv2.putText(frame, str(face_percent) + "%", (left+6,bottom+26), font, 0.6, (255,255,255), 1)
            
        
        cv2.imshow("FaceRecognition by M6/3", frame)
        if cv2.waitKey(1) & 0xFF == ord("q"):
            break
    nameofpc = os.getlogin()
    excel_path = os.path.realpath('ExcelData')
    workbook.save(f'{excel_path}/{dtt}.xlsx')

    cv2.destroyAllWindows()

if __name__ =='__main__':

    print('Run program on gui.py file I-SUS.')
